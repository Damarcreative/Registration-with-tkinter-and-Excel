import tkinter as tk
import openpyxl

# Create a registration form
form = tk.Tk()
form.title("Student Registration")

# Create labels and entries for names
label_name = tk.Label(form, text="Name")
entry_name = tk.Entry(form)
label_name.grid(row=0, column=0)
entry_name.grid(row=0, column=1)

# Create labels and entries for addresses
label_address = tk.Label(form, text="Address")
entry_address = tk.Entry(form)
label_address.grid(row=1, column=0)
entry_address.grid(row=1, column=1)

# Create labels and entries for phone numbers
label_telp = tk.Label(form, text="Phone number")
entry_telp = tk.Entry(form)
label_telp.grid(row=2, column=0)
entry_telp.grid(row=2, column=1)


# Create a function to store data
def send():
  # Open excel file
  wb = openpyxl.load_workbook("data_siswa.xlsx")
  # Choose the sheet to use
  sheet = wb["Sheet1"]

  # Specifies the location of the data to be stored
  row = sheet.max_row + 1
  col = 1

  # Save data into excel files
  sheet.cell(row=row, column=col).value = entry_name.get()
  sheet.cell(row=row, column=col+1).value = entry_address.get()
  sheet.cell(row=row, column=col+2).value = entry_telp.get()

  # Save changes
  wb.save("data_siswa.xlsx")

  # Displays a success message
  tk.Label(form, text="Stored data!").grid(row=4, column=0)

# Create a button to save data
btn_simpan = tk.Button(form, text="Send", command=send)
btn_simpan.grid(row=3, column=0)

form.mainloop()
